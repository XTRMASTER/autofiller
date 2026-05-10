import openpyxl
import os

class ExcelHandler:
    @staticmethod
    def scan_excel_document(filepath: str):
        """
        Scans an Excel document and returns cell coordinates and values.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str) and str(cell.value).strip():
                        content.append({
                            'sheet': sheet_name,
                            'address': cell.coordinate,
                            'text': str(cell.value)
                        })
        wb.close()
        return content

    @staticmethod
    def apply_excel_variables(filepath: str, variables_dict: dict, output_path: str):
        """
        Replaces variables in an Excel document and saves to output_path.
        """
        wb = openpyxl.load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        text = str(cell.value)
                        changed = False
                        for old_val, new_val in variables_dict.items():
                            if old_val in text:
                                text = text.replace(old_val, new_val)
                                changed = True
                        if changed:
                            cell.value = text
        wb.save(output_path)
        wb.close()
